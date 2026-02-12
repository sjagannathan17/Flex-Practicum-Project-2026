"""
Excel export functionality using openpyxl.
Generates comprehensive Excel reports with multiple sheets.
"""
import io
from datetime import datetime
from typing import Optional

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    from openpyxl.chart import BarChart, PieChart, Reference
    from openpyxl.utils.dataframe import dataframe_to_rows
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

from backend.core.config import COMPANIES
from backend.analytics.sentiment import analyze_company_sentiment
from backend.analytics.classifier import classify_company_investments, compare_investment_focus
from backend.analytics.trends import analyze_company_trends, compare_company_trends
from backend.analytics.geographic import get_company_facilities, compare_geographic_footprints
from backend.analytics.anomaly import get_all_anomalies


def _apply_header_style(cell):
    """Apply header styling to a cell."""
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center")


def _apply_data_style(cell, row_idx):
    """Apply alternating row colors."""
    if row_idx % 2 == 0:
        cell.fill = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")


def generate_excel_report(company: str) -> bytes:
    """
    Generate a comprehensive Excel report for a single company.
    
    Returns:
        Excel file as bytes
    """
    if not HAS_OPENPYXL:
        raise ImportError("openpyxl is required for Excel export. Install with: pip install openpyxl")
    
    company_title = company.title()
    wb = Workbook()
    
    # === Overview Sheet ===
    ws_overview = wb.active
    ws_overview.title = "Overview"
    
    # Title
    ws_overview["A1"] = f"{company_title} - Competitive Intelligence Report"
    ws_overview["A1"].font = Font(size=18, bold=True)
    ws_overview.merge_cells("A1:E1")
    
    ws_overview["A2"] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws_overview["A2"].font = Font(italic=True, color="666666")
    
    # Get data
    sentiment = analyze_company_sentiment(company_title)
    classification = classify_company_investments(company_title)
    trends = analyze_company_trends(company_title)
    facilities = get_company_facilities(company_title)
    
    # Key Metrics
    headers = ["Metric", "Value", "Details"]
    for col, header in enumerate(headers, 1):
        cell = ws_overview.cell(row=4, column=col, value=header)
        _apply_header_style(cell)
    
    metrics = [
        ("Sentiment Score", f"{sentiment.get('sentiment_score', 0):.1%}", "Based on document analysis"),
        ("AI Focus", f"{classification.get('overall_ai_focus_percentage', 0):.1f}%", classification.get('investment_focus', 'N/A')),
        ("Trend Outlook", trends.get('overall_outlook', 'N/A').title(), "Based on CapEx, AI focus, sentiment"),
        ("Total Facilities", facilities.get('total_facilities', 0), f"HQ: {facilities.get('headquarters', {}).get('city', 'N/A')}"),
        ("Documents Analyzed", classification.get('total_documents', 0), "SEC filings and transcripts"),
    ]
    
    for row_idx, (metric, value, details) in enumerate(metrics, 5):
        ws_overview.cell(row=row_idx, column=1, value=metric)
        ws_overview.cell(row=row_idx, column=2, value=str(value))
        ws_overview.cell(row=row_idx, column=3, value=details)
        for col in range(1, 4):
            _apply_data_style(ws_overview.cell(row=row_idx, column=col), row_idx)
    
    # Adjust column widths
    ws_overview.column_dimensions['A'].width = 25
    ws_overview.column_dimensions['B'].width = 20
    ws_overview.column_dimensions['C'].width = 40
    
    # === Investment Analysis Sheet ===
    ws_invest = wb.create_sheet("Investment Analysis")
    
    ws_invest["A1"] = "Investment Classification"
    ws_invest["A1"].font = Font(size=14, bold=True)
    
    headers = ["Category", "Document Count", "Percentage"]
    for col, header in enumerate(headers, 1):
        cell = ws_invest.cell(row=3, column=col, value=header)
        _apply_header_style(cell)
    
    breakdown = classification.get('investment_breakdown', {})
    row = 4
    for category, data in breakdown.items():
        ws_invest.cell(row=row, column=1, value=category.replace('_', ' ').title())
        ws_invest.cell(row=row, column=2, value=data.get('count', 0))
        ws_invest.cell(row=row, column=3, value=f"{data.get('percentage', 0):.1f}%")
        for col in range(1, 4):
            _apply_data_style(ws_invest.cell(row=row, column=col), row)
        row += 1
    
    ws_invest.column_dimensions['A'].width = 20
    ws_invest.column_dimensions['B'].width = 18
    ws_invest.column_dimensions['C'].width = 15
    
    # === Trend Analysis Sheet ===
    ws_trends = wb.create_sheet("Trends")
    
    ws_trends["A1"] = "Trend Analysis"
    ws_trends["A1"].font = Font(size=14, bold=True)
    
    headers = ["Metric", "Direction", "Confidence", "Forecast"]
    for col, header in enumerate(headers, 1):
        cell = ws_trends.cell(row=3, column=col, value=header)
        _apply_header_style(cell)
    
    trend_data = [
        ("CapEx", trends.get('capex_trend', {})),
        ("AI Focus", trends.get('ai_focus_trend', {})),
        ("Sentiment", trends.get('sentiment_trend', {})),
    ]
    
    row = 4
    for metric, data in trend_data:
        ws_trends.cell(row=row, column=1, value=metric)
        ws_trends.cell(row=row, column=2, value=data.get('direction', 'N/A').title())
        ws_trends.cell(row=row, column=3, value=f"{data.get('confidence', 0):.1f}%")
        forecast = data.get('next_period_forecast')
        ws_trends.cell(row=row, column=4, value=f"{forecast:.2f}" if forecast else "N/A")
        for col in range(1, 5):
            _apply_data_style(ws_trends.cell(row=row, column=col), row)
        row += 1
    
    ws_trends.column_dimensions['A'].width = 15
    ws_trends.column_dimensions['B'].width = 15
    ws_trends.column_dimensions['C'].width = 15
    ws_trends.column_dimensions['D'].width = 15
    
    # === Facilities Sheet ===
    ws_facilities = wb.create_sheet("Facilities")
    
    ws_facilities["A1"] = "Global Facilities"
    ws_facilities["A1"].font = Font(size=14, bold=True)
    
    headers = ["City", "Country", "Region", "Type"]
    for col, header in enumerate(headers, 1):
        cell = ws_facilities.cell(row=3, column=col, value=header)
        _apply_header_style(cell)
    
    # Add headquarters
    hq = facilities.get('headquarters', {})
    if hq:
        ws_facilities.cell(row=4, column=1, value=hq.get('city', ''))
        ws_facilities.cell(row=4, column=2, value=hq.get('country', ''))
        ws_facilities.cell(row=4, column=3, value=hq.get('region', ''))
        ws_facilities.cell(row=4, column=4, value='Headquarters')
        for col in range(1, 5):
            cell = ws_facilities.cell(row=4, column=col)
            cell.fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
    
    row = 5
    for facility in facilities.get('facilities', []):
        ws_facilities.cell(row=row, column=1, value=facility.get('city', ''))
        ws_facilities.cell(row=row, column=2, value=facility.get('country', ''))
        ws_facilities.cell(row=row, column=3, value=facility.get('region', ''))
        ws_facilities.cell(row=row, column=4, value=facility.get('type', 'Manufacturing'))
        for col in range(1, 5):
            _apply_data_style(ws_facilities.cell(row=row, column=col), row)
        row += 1
    
    for col in ['A', 'B', 'C', 'D']:
        ws_facilities.column_dimensions[col].width = 20
    
    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def generate_comparison_excel() -> bytes:
    """
    Generate an Excel report comparing all companies.
    
    Returns:
        Excel file as bytes
    """
    if not HAS_OPENPYXL:
        raise ImportError("openpyxl is required for Excel export. Install with: pip install openpyxl")
    
    wb = Workbook()
    
    # === Comparison Summary Sheet ===
    ws = wb.active
    ws.title = "Comparison Summary"
    
    ws["A1"] = "EMS Industry Competitive Analysis"
    ws["A1"].font = Font(size=18, bold=True)
    ws.merge_cells("A1:G1")
    
    ws["A2"] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws["A2"].font = Font(italic=True, color="666666")
    
    # Get comparison data
    sentiment_comparison = {"companies": {c['name'].split()[0]: analyze_company_sentiment(c['name'].split()[0]) for c in COMPANIES.values()}}
    investment_comparison = compare_investment_focus()
    trends_comparison = compare_company_trends()
    geo_comparison = compare_geographic_footprints()
    
    # Headers
    headers = ["Company", "Sentiment", "AI Focus %", "Trend Outlook", "Facilities", "Primary Region"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        _apply_header_style(cell)
    
    row = 5
    for ticker, config in COMPANIES.items():
        company_name = config['name'].split()[0]
        
        sentiment = sentiment_comparison.get('companies', {}).get(company_name, {})
        investment = investment_comparison.get('companies', {}).get(company_name, {})
        trend = trends_comparison.get('companies', {}).get(company_name, {})
        geo = geo_comparison.get('companies', {}).get(company_name, {})
        
        ws.cell(row=row, column=1, value=company_name)
        ws.cell(row=row, column=2, value=f"{sentiment.get('sentiment_score', 0):.1%}")
        ws.cell(row=row, column=3, value=f"{investment.get('ai_focus_percentage', 0):.1f}%")
        ws.cell(row=row, column=4, value=trend.get('outlook', 'N/A').title())
        ws.cell(row=row, column=5, value=geo.get('total_facilities', 0))
        ws.cell(row=row, column=6, value=geo.get('primary_region', 'N/A'))
        
        for col in range(1, 7):
            _apply_data_style(ws.cell(row=row, column=col), row)
        row += 1
    
    # Adjust widths
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 18
    
    # === Anomalies Sheet ===
    ws_anomalies = wb.create_sheet("Anomalies")
    
    ws_anomalies["A1"] = "Detected Anomalies"
    ws_anomalies["A1"].font = Font(size=14, bold=True)
    
    headers = ["Company", "Type", "Severity", "Period", "Change %"]
    for col, header in enumerate(headers, 1):
        cell = ws_anomalies.cell(row=3, column=col, value=header)
        _apply_header_style(cell)
    
    anomalies = get_all_anomalies()
    row = 4
    for company, company_anomalies in anomalies.items():
        for anomaly_type, anomaly_list in company_anomalies.items():
            if isinstance(anomaly_list, list):
                for anomaly in anomaly_list:
                    ws_anomalies.cell(row=row, column=1, value=company)
                    ws_anomalies.cell(row=row, column=2, value=anomaly_type.replace('_', ' ').title())
                    ws_anomalies.cell(row=row, column=3, value=anomaly.get('severity', 'medium').title())
                    ws_anomalies.cell(row=row, column=4, value=anomaly.get('period', 'N/A'))
                    ws_anomalies.cell(row=row, column=5, value=f"{anomaly.get('pct_change_from_mean', 0):.1f}%")
                    for col in range(1, 6):
                        _apply_data_style(ws_anomalies.cell(row=row, column=col), row)
                    row += 1
    
    for col in ['A', 'B', 'C', 'D', 'E']:
        ws_anomalies.column_dimensions[col].width = 18
    
    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()
